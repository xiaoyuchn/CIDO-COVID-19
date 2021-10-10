from rdflib import Graph, Literal, URIRef
from openpyxl import load_workbook

g=Graph()
g.parse("cido-covid-19.owl",format="xml")

label=URIRef("http://www.w3.org/2000/01/rdf-schema#label")
comment=URIRef("http://www.w3.org/2000/01/rdf-schema#comment")
definition=URIRef("http://purl.obolibrary.org/obo/IAO_0000115")
label_en=URIRef("http://purl.obolibrary.org/obo/cido-covid-19.owl#label_en")

filename = r'cido-covid-19_info_relation - zh.xlsx'
workbook = load_workbook(filename, read_only=False, data_only=True)
ws = workbook.active

for row in ws.iter_rows(min_row=2, min_col=1, max_col=7, max_row=453, values_only=True):
    if row[1]!=row[2]:
        g.remove((URIRef(row[0]), label, None))
        g.add((URIRef(row[0]), label_en, Literal(row[1])))
        g.add((URIRef(row[0]), label, Literal(row[2])))
    if row[3]!=row[4]:
        labelword = Literal(row[4])
        g.add((URIRef(row[0]), comment, labelword))
    if row[5]!=row[6]:
        labelword = Literal(row[6])
        g.add((URIRef(row[0]), definition, labelword))

g.serialize(destination="cido-covid-19.owl")

print("ok")




    

    
