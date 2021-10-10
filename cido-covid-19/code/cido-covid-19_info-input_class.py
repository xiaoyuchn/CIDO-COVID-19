from rdflib import Graph, Literal, URIRef
from openpyxl import load_workbook

g=Graph()
g.parse("cido-covid-19.owl",format="xml")

label=URIRef("http://www.w3.org/2000/01/rdf-schema#label")
comment=URIRef("http://www.w3.org/2000/01/rdf-schema#comment")
definition=URIRef("http://purl.obolibrary.org/obo/IAO_0000115")
label_en=URIRef("http://purl.obolibrary.org/obo/cido-covid-19.owl#label_en")

filename = r'C:\Users\xiaoyu\OneDrive\xiaoyu\课题119\python代码\owl处理\cido-covid-19_info_class - en - 副本.xlsx'
workbook = load_workbook(filename, read_only=False, data_only=True)
ws = workbook.active

for row in ws.iter_rows(min_row=3, min_col=1, max_col=5, max_row=10342, values_only=True):
    if row[1]!=row[2]:
        g.remove((URIRef(row[0]), label, None))
        g.add((URIRef(row[0]), label_en, Literal(row[1])))
        g.add((URIRef(row[0]), label, Literal(row[2])))
    if row[3]!=row[4]:
        labelword = Literal(row[4])
        g.add((URIRef(row[0]), definition, labelword))

g.serialize(destination="cido-covid-19.owl")

print("完事")




    

    
