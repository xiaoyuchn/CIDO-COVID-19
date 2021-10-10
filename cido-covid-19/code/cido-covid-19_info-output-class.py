from rdflib import Graph, URIRef
import openpyxl

g = Graph()
g.parse(r"C:\Users\xiaoyu\OneDrive\xiaoyu\课题119\cido-covid-19\cido-covid-19.owl", format="xml")

i = 1
wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0)

sub = URIRef("http://www.w3.org/2000/01/rdf-schema#subClassOf")  # 注意是class还是property
XX = URIRef("http://purl.obolibrary.org/obo/BFO_0000001")   #最顶层术语，此处为“entity”

ll = set()

def bianli1(n):
    global i
    for s, p, o in g.triples((None, sub, n)):
        for S, P, O in g.triples((s, None, None)):
            ll.add(P)
        XX = s
        bianli1(XX)

for s, p, o in g.triples((XX, None, None)):
    ll.add(p)
bianli1(XX)
ls = list(ll)

def bianli(n):
    global i
    for s, p, o in g.triples((None, sub, n)):
        ws.cell(i + 2, 1, 5)
        ws.cell(i + 2, 2, g.label(s))
        ws.cell(i + 2, 3, s)
        for j in range(len(ls)):
            YY = ls[j]
            var = ""
            for a, b, c in g.triples((s, YY, None)):
                var += c
                var += ";"
            ws.cell(i + 2, j + 4, var)
        i += 1
        XX = s
        bianli(XX)

def main():
    ws.cell(1, 1, "class_id")
    ws.cell(1, 2, "class_name")
    ws.cell(1, 3, "class_url")
    for n in range(len(ls)):
        ws.cell(1, n + 4, ls[n])
    ws.cell(2, 1, i)
    ws.cell(2, 2, g.label(XX))
    ws.cell(2, 3, XX)
    bianli(XX)
    savexlsx = "cido-covid-19_info_class.xlsx"
    wb.save(savexlsx)

main()
